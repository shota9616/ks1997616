#!/usr/bin/env python3
"""その他9種書類生成"""

import os
import shutil
from pathlib import Path

import openpyxl
from openpyxl.cell.cell import MergedCell
from docx import Document

from models import HearingData
from config import Config
from hearing_reader import _split_name, _find_sheet_in_workbook


def generate_other_documents(data: HearingData, output_dir: str, template_dir: Path):
    """その他の書類を生成（openpyxlでデータ入力）"""
    print("\n📄 その他の書類を生成中...")

    c = data.company
    f = data.funding

    def safe_write(ws, cell_addr, value):
        """マージセルでも安全に書き込む"""
        cell = ws[cell_addr]
        if isinstance(cell, MergedCell):
            for mc in ws.merged_cells.ranges:
                if cell.coordinate in mc:
                    ws.cell(mc.min_row, mc.min_col).value = value
                    return
        else:
            cell.value = value

    # === 1. 役員名簿 ===
    try:
        src = template_dir / "役員名簿_様式.xlsx"
        dst = Path(output_dir) / "役員名簿_完成版.xlsx"
        shutil.copy(src, dst)
        wb = openpyxl.load_workbook(dst)
        # Phase 6: ファジーマッチング
        ws = _find_sheet_in_workbook(wb, ["役員名簿", "役員"]) or wb[wb.sheetnames[0]]
        safe_write(ws, 'D5', c.name)  # 法人名
        safe_write(ws, 'D7', c.officer_count)  # 役員数
        safe_write(ws, 'D8', 0)  # 大企業所属人数
        safe_write(ws, 'D9', 0)  # みなし大企業所属人数
        for i, off in enumerate(data.officers[:10]):
            row = 15 + i
            ws[f'B{row}'] = i + 1
            ws[f'C{row}'] = off.position
            # Phase 6: _split_name で全角スペース対応
            last_name, first_name = _split_name(off.name)
            ws[f'D{row}'] = last_name
            ws[f'E{row}'] = first_name
        wb.save(dst)
        wb.close()
        print(f"    ✅ 役員名簿_完成版.xlsx（{c.officer_count}名）")
    except Exception as e:
        print(f"    ⚠️ 役員名簿エラー: {e}")

    # === 2. 従業員名簿 ===
    try:
        src = template_dir / "従業員名簿_様式.xlsx"
        dst = Path(output_dir) / "従業員名簿_完成版.xlsx"
        shutil.copy(src, dst)
        wb = openpyxl.load_workbook(dst)
        # Phase 6: ファジーマッチング
        ws = _find_sheet_in_workbook(wb, ["労働者名簿", "従業員名簿", "従業員"]) or wb[wb.sheetnames[0]]
        safe_write(ws, 'C5', c.name)
        safe_write(ws, 'C7', c.employee_count)
        for i, emp in enumerate(data.employees[:50]):
            row = 12 + i
            ws[f'B{row}'] = i + 1
            # Phase 6: _split_name で全角スペース対応
            last_name, first_name = _split_name(emp.name)
            ws[f'C{row}'] = last_name
            ws[f'D{row}'] = first_name
            if emp.birth_date:
                ws[f'E{row}'] = emp.birth_date
        wb.save(dst)
        wb.close()
        print(f"    ✅ 従業員名簿_完成版.xlsx（{c.employee_count}名）")
    except Exception as e:
        print(f"    ⚠️ 従業員名簿エラー: {e}")

    # === 3. 株主・出資者名簿 ===
    try:
        src = template_dir / "株主出資者名簿_様式.xlsx"
        dst = Path(output_dir) / "株主出資者名簿_完成版.xlsx"
        shutil.copy(src, dst)
        wb = openpyxl.load_workbook(dst)
        # Phase 6: ファジーマッチング
        ws = _find_sheet_in_workbook(wb, ["株主名簿", "株主"]) or wb[wb.sheetnames[0]]
        safe_write(ws, 'C5', c.name)
        safe_write(ws, 'C6', c.capital)
        for i, sh in enumerate(data.shareholders[:20]):
            row = 14 + i
            ws[f'B{row}'] = i + 1
            ws[f'C{row}'] = sh.name
            ws[f'D{row}'] = sh.shares
        wb.save(dst)
        wb.close()
        print(f"    ✅ 株主出資者名簿_完成版.xlsx（{len(data.shareholders)}名）")
    except Exception as e:
        print(f"    ⚠️ 株主名簿エラー: {e}")

    # === 4. 事業実施場所リスト ===
    try:
        src = template_dir / "事業実施場所リスト_様式.xlsx"
        dst = Path(output_dir) / "事業実施場所リスト_完成版.xlsx"
        shutil.copy(src, dst)
        wb = openpyxl.load_workbook(dst)
        # Phase 6: ファジーマッチング
        ws = _find_sheet_in_workbook(wb, ["所在地リスト", "所在地", "事業実施場所"]) or wb[wb.sheetnames[0]]
        ws['C20'] = c.postal_code.replace('-', '') if c.postal_code else ''
        ws['C21'] = c.prefecture
        addr_parts = c.address.replace(c.prefecture, '').strip() if c.address else ''
        ws['C22'] = addr_parts[:10] if addr_parts else ''
        ws['C23'] = addr_parts[10:20] if len(addr_parts) > 10 else ''
        ws['C24'] = addr_parts[20:] if len(addr_parts) > 20 else ''
        ws['C26'] = c.name + ' 本社'
        ws['C27'] = c.phone
        wb.save(dst)
        wb.close()
        print(f"    ✅ 事業実施場所リスト_完成版.xlsx")
    except Exception as e:
        print(f"    ⚠️ 事業実施場所リストエラー: {e}")

    # === 5. 他の補助金使用実績 ===
    try:
        src = template_dir / "他の補助金使用実績_様式.xlsx"
        dst = Path(output_dir) / "他の補助金使用実績_完成版.xlsx"
        shutil.copy(src, dst)
        wb = openpyxl.load_workbook(dst)
        # Phase 6: ファジーマッチング
        ws = _find_sheet_in_workbook(wb, ["Sheet1", "他の補助金", "使用実績"]) or wb[wb.sheetnames[0]]
        ws['C25'] = "なし"
        ws['C26'] = "-"
        ws['C27'] = "-"
        ws['C28'] = "-"
        wb.save(dst)
        wb.close()
        print(f"    ✅ 他の補助金使用実績_完成版.xlsx")
    except Exception as e:
        print(f"    ⚠️ 他の補助金使用実績エラー: {e}")

    # === 6. 給与支給総額確認書 ===
    try:
        src = template_dir / "給与支給総額確認書_様式.xlsx"
        dst = Path(output_dir) / "給与支給総額確認書_完成版.xlsx"
        shutil.copy(src, dst)
        wb = openpyxl.load_workbook(dst)
        if "宣誓書" in wb.sheetnames:
            ws = wb["宣誓書"]
            safe_write(ws, 'C8', c.name)
            safe_write(ws, 'E8', c.representative)
        for sname in wb.sheetnames:
            if "直近決算" in sname and "記入例" not in sname and "未満" not in sname:
                ws = wb[sname]
                safe_write(ws, 'C5', c.name)
                base_salary = c.total_salary if c.total_salary > 0 else int(c.revenue_2024 * Config.SALARY_RATIO)
                safe_write(ws, 'E10', base_salary)
                safe_write(ws, 'E11', c.employee_count)
                break
        wb.save(dst)
        wb.close()
        print(f"    ✅ 給与支給総額確認書_完成版.xlsx")
    except Exception as e:
        print(f"    ⚠️ 給与支給総額確認書エラー: {e}")

    # === 7. 賃金引上げ要件（事業場内） ===
    try:
        src = template_dir / "賃金引上げ要件_事業場内_様式.xlsx"
        dst = Path(output_dir) / "賃金引上げ要件_事業場内_完成版.xlsx"
        shutil.copy(src, dst)
        wb = openpyxl.load_workbook(dst)
        if "確認書" in wb.sheetnames:
            ws = wb["確認書"]
            safe_write(ws, 'C7', c.name)
            safe_write(ws, 'E7', c.representative)
        wb.save(dst)
        wb.close()
        print(f"    ✅ 賃金引上げ要件_事業場内_完成版.xlsx")
    except Exception as e:
        print(f"    ⚠️ 賃金引上げ要件_事業場内エラー: {e}")

    # === 8. 賃金引上げ要件（地域別） ===
    try:
        src = template_dir / "賃金引上げ要件_地域別_様式.xlsx"
        dst = Path(output_dir) / "賃金引上げ要件_地域別_完成版.xlsx"
        shutil.copy(src, dst)
        wb = openpyxl.load_workbook(dst)
        if "確認書" in wb.sheetnames:
            ws = wb["確認書"]
            safe_write(ws, 'C7', c.name)
            safe_write(ws, 'E7', c.representative)
        wb.save(dst)
        wb.close()
        print(f"    ✅ 賃金引上げ要件_地域別_完成版.xlsx")
    except Exception as e:
        print(f"    ⚠️ 賃金引上げ要件_地域別エラー: {e}")

    # === 9. 金融機関確認書（Word） ===
    try:
        src = template_dir / "金融機関確認書_様式.docx"
        dst = Path(output_dir) / "金融機関確認書_完成版.docx"
        shutil.copy(src, dst)
        doc = Document(str(dst))
        for table in doc.tables:
            for row in table.rows:
                for i, cell in enumerate(row.cells):
                    txt = cell.text
                    if '法人名' in txt or '申請者名' in txt or '事業者名' in txt:
                        if i + 1 < len(row.cells):
                            row.cells[i + 1].text = c.name
                    if '代表者名' in txt:
                        if i + 1 < len(row.cells):
                            row.cells[i + 1].text = c.representative
                    if '金融機関名' in txt:
                        if i + 1 < len(row.cells):
                            row.cells[i + 1].text = f.bank_name
        doc.save(str(dst))
        print(f"    ✅ 金融機関確認書_完成版.docx")
    except Exception as e:
        print(f"    ⚠️ 金融機関確認書エラー: {e}")
