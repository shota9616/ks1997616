#!/usr/bin/env python3
"""
議事録テキスト → ヒアリングシート自動生成パイプライン

ミーティングの文字起こしテキストからClaude APIで構造化データを抽出し、
ヒアリングシートExcel（11シート）を自動生成する。

【使用方法】
  python scripts/transcription_to_hearing.py --input 議事録.txt --output hearing.xlsx
  python scripts/transcription_to_hearing.py --input 議事録.txt --output hearing.xlsx --generate
"""

import argparse
import json
import os
import re
import sys
import tempfile
from dataclasses import dataclass, field
from pathlib import Path
from typing import List, Dict, Optional

import openpyxl

# scripts/ を import path に追加
sys.path.insert(0, str(Path(__file__).parent))

from models import (
    HearingData,
    CompanyInfo,
    LaborShortageInfo,
    LaborSavingInfo,
    EquipmentInfo,
    FundingInfo,
    OfficerInfo,
    EmployeeInfo,
    ShareholderInfo,
    WorkProcess,
)
from config import Config
from hearing_reader import read_hearing_sheet

# anthropic SDK
try:
    import anthropic
    ANTHROPIC_AVAILABLE = True
except ImportError:
    ANTHROPIC_AVAILABLE = False


# =============================================================================
# 日本語数値パーサー
# =============================================================================

def _normalize_japanese_number(text: str) -> int:
    """日本語数値表記を整数に変換する。

    Examples:
        "480万円"      → 4800000
        "1億2000万"    → 120000000
        "約500万"      → 5000000
        "3,000万円"    → 30000000
        "1200"         → 1200
        "12000000"     → 12000000
    """
    if not text:
        return 0
    text = str(text).strip()
    # 概数マーカー・通貨記号等を除去
    text = re.sub(r'[約およそほぼ円￥¥、,\s]', '', text)
    if not text:
        return 0

    # 「億」と「万」の処理
    oku = 0
    man = 0
    remainder = 0

    m_oku = re.search(r'([\d.]+)\s*億', text)
    if m_oku:
        oku = float(m_oku.group(1)) * 100_000_000
        text = text[:m_oku.start()] + text[m_oku.end():]

    m_man = re.search(r'([\d.]+)\s*万', text)
    if m_man:
        man = float(m_man.group(1)) * 10_000
        text = text[:m_man.start()] + text[m_man.end():]

    # 残りの数値
    text = text.strip()
    if text:
        try:
            remainder = float(text)
        except ValueError:
            pass

    total = oku + man + remainder
    return int(total)


# =============================================================================
# Claude API JSON抽出ヘルパー
# =============================================================================

def _parse_json_from_response(text: str) -> dict:
    """Claude APIレスポンスからJSONを抽出する（pdf_extractor.pyパターン踏襲）"""
    cleaned = text.strip()
    # ```json ... ``` マーカーを除去
    if cleaned.startswith("```"):
        lines = cleaned.split("\n")
        lines = [l for l in lines if not l.strip().startswith("```")]
        cleaned = "\n".join(lines)
    # JSON部分を抽出
    start = cleaned.find("{")
    end = cleaned.rfind("}") + 1
    if start >= 0 and end > start:
        try:
            return json.loads(cleaned[start:end])
        except json.JSONDecodeError:
            pass
    return {}


def _call_claude(client, text: str, prompt: str, model: str) -> dict:
    """Claude APIにテキストを送信してJSONを取得する"""
    message = client.messages.create(
        model=model,
        max_tokens=4096,
        messages=[
            {
                "role": "user",
                "content": f"{prompt}\n\n---\n以下が議事録テキストです:\n\n{text}",
            }
        ],
    )
    response_text = message.content[0].text
    return _parse_json_from_response(response_text)


# =============================================================================
# Claude API 抽出層（4回の専門別抽出）
# =============================================================================

PROMPT_COMPANY_BASICS = """以下の議事録テキストから、企業基本情報・人手不足の実態・導入の動機に関する情報を抽出してください。
必ず以下のJSON形式で返してください。JSONのみ返し、説明文は不要です。
見つからないフィールドは空文字""または0としてください。
金額は円単位の整数にしてください（「480万円」→4800000）。

{
    "company_name": "",
    "representative": "",
    "prefecture": "",
    "address": "",
    "postal_code": "",
    "phone": "",
    "established_date": "",
    "capital": 0,
    "industry": "",
    "business_description": "",
    "employee_count": 0,
    "officer_count": 1,
    "url": "",
    "shortage_tasks": "",
    "recruitment_period": "",
    "applications": 0,
    "hired": 0,
    "overtime_hours": 0,
    "current_workers": 0,
    "desired_workers": 0,
    "job_openings_ratio": 0,
    "motivation_background": ""
}"""

PROMPT_EQUIPMENT_FUNDING = """以下の議事録テキストから、導入する設備・資金調達に関する情報を抽出してください。
必ず以下のJSON形式で返してください。JSONのみ返し、説明文は不要です。
見つからないフィールドは空文字""または0としてください。
金額は円単位の整数にしてください（「480万円」→4800000）。

{
    "equipment_name": "",
    "equipment_category": "",
    "manufacturer": "",
    "model": "",
    "quantity": 1,
    "total_price": 0,
    "vendor": "",
    "features": "",
    "catalog_number": "",
    "total_investment": 0,
    "subsidy_amount": 0,
    "self_funding": 0,
    "implementation_manager": "",
    "implementation_period": "",
    "bank_name": ""
}"""

PROMPT_EMPLOYEES = """以下の議事録テキストから、従業員・役員・株主に関する情報を抽出してください。
必ず以下のJSON形式で返してください。JSONのみ返し、説明文は不要です。
見つからないフィールドは空文字""または0としてください。

{
    "officers": [
        {"name": "", "position": "", "birth_date": ""}
    ],
    "employees": [
        {"name": "", "birth_date": "", "hire_date": ""}
    ],
    "shareholders": [
        {"name": "", "shares": 0}
    ]
}

【注意】
- 議事録に名前が出てこない場合は空の配列[]を返してください
- 代表者は役員にも含めてください
- 役職は「代表取締役」「取締役」「監査役」等"""

PROMPT_EFFECTS_FINANCE = """以下の議事録テキストから、省力化効果・財務情報・賃上げ計画・活用計画に関する情報を抽出してください。
必ず以下のJSON形式で返してください。JSONのみ返し、説明文は不要です。
見つからないフィールドは空文字""または0としてください。
金額は円単位の整数にしてください（「480万円」→4800000）。
時間は小数で返してください（「1時間30分」→1.5）。

{
    "target_tasks": "",
    "current_hours": 0,
    "target_hours": 0,
    "revenue": 0,
    "gross_profit": 0,
    "operating_profit": 0,
    "labor_cost": 0,
    "depreciation": 0,
    "total_salary": 0,
    "wage_increase_rate": 0,
    "wage_increase_target": "",
    "wage_increase_timing": "",
    "time_utilization_plan": ""
}"""


def extract_from_transcription(
    text: str, api_key: str, model: str = "claude-sonnet-4-20250514"
) -> dict:
    """議事録テキストから4回のClaude API呼び出しで全フィールドを抽出する。

    Returns:
        統合された辞書（全4回の結果をマージ）
    """
    if not ANTHROPIC_AVAILABLE:
        raise ImportError("anthropic パッケージがインストールされていません。pip install anthropic を実行してください。")

    client = anthropic.Anthropic(api_key=api_key)
    merged = {}

    calls = [
        ("企業基本情報+人手不足+動機", PROMPT_COMPANY_BASICS),
        ("導入設備+資金調達", PROMPT_EQUIPMENT_FUNDING),
        ("従業員+役員+株主", PROMPT_EMPLOYEES),
        ("省力化効果+財務+賃上げ", PROMPT_EFFECTS_FINANCE),
    ]

    for label, prompt in calls:
        print(f"  📡 Claude API呼び出し: {label}")
        result = _call_claude(client, text, prompt, model)
        if result:
            merged.update(result)
            print(f"    ✅ {len(result)}フィールド抽出")
        else:
            print(f"    ⚠️ 抽出結果が空でした")

    return merged


# =============================================================================
# バリデーション層
# =============================================================================

@dataclass
class ExtractionResult:
    """抽出結果 + 信頼度情報"""
    data: dict = field(default_factory=dict)
    warnings: List[str] = field(default_factory=list)
    missing_critical: List[str] = field(default_factory=list)
    missing_optional: List[str] = field(default_factory=list)


# 重要フィールド（欠落時は警告）
_CRITICAL_FIELDS = {
    "company_name": "会社名",
    "industry": "業種",
    "equipment_name": "設備名",
    "total_price": "設備金額",
    "employee_count": "従業員数",
}

# オプションフィールド（欠落時は情報提供のみ）
_OPTIONAL_FIELDS = {
    "url": "ホームページURL",
    "catalog_number": "カタログ番号",
    "phone": "電話番号",
    "postal_code": "郵便番号",
    "bank_name": "取引銀行",
}


def validate_extracted_data(raw: dict) -> ExtractionResult:
    """抽出データをバリデーションし、ExtractionResultを返す。"""
    result = ExtractionResult(data=raw)

    # 重要フィールドチェック
    for key, label in _CRITICAL_FIELDS.items():
        val = raw.get(key)
        if not val or (isinstance(val, (int, float)) and val == 0):
            result.missing_critical.append(label)

    # オプションフィールドチェック
    for key, label in _OPTIONAL_FIELDS.items():
        val = raw.get(key)
        if not val or (isinstance(val, (int, float)) and val == 0):
            result.missing_optional.append(label)

    # 数値の妥当性チェック
    total_price = raw.get("total_price", 0)
    if isinstance(total_price, str):
        total_price = _normalize_japanese_number(total_price)
        raw["total_price"] = total_price
    if isinstance(total_price, (int, float)) and total_price < 0:
        result.warnings.append(f"設備金額が負数です: {total_price}")

    employee_count = raw.get("employee_count", 0)
    if isinstance(employee_count, (int, float)) and employee_count > 1000:
        result.warnings.append(f"従業員数が1000名を超えています: {employee_count}（中小企業要件を確認してください）")

    current_hours = raw.get("current_hours", 0)
    target_hours = raw.get("target_hours", 0)
    if isinstance(current_hours, (int, float)) and isinstance(target_hours, (int, float)):
        if 0 < current_hours < target_hours:
            result.warnings.append(f"導入後の時間({target_hours}h)が導入前({current_hours}h)より大きいです")

    # 警告サマリ
    if result.missing_critical:
        result.warnings.insert(0, f"重要フィールド未検出: {', '.join(result.missing_critical)}")
    if result.missing_optional:
        result.warnings.append(f"任意フィールド未検出: {', '.join(result.missing_optional)}")

    return result


# =============================================================================
# HearingData 変換
# =============================================================================

def _safe_int(val, default=0) -> int:
    """値を安全にintに変換する。日本語数値もパースする。"""
    if val is None:
        return default
    if isinstance(val, str):
        parsed = _normalize_japanese_number(val)
        return parsed if parsed else default
    try:
        return int(float(val))
    except (ValueError, TypeError):
        return default


def _safe_float(val, default=0.0) -> float:
    """値を安全にfloatに変換する。"""
    if val is None:
        return default
    if isinstance(val, str):
        val = re.sub(r'[時間hH%％]', '', val).strip()
        if not val:
            return default
    try:
        return float(val)
    except (ValueError, TypeError):
        return default


def _safe_str(val, default="") -> str:
    """値を安全にstrに変換する。"""
    if val is None:
        return default
    return str(val).strip()


def build_hearing_data(result: ExtractionResult) -> HearingData:
    """ExtractionResultから既存HearingDataデータクラスにマッピングする。"""
    d = result.data
    data = HearingData()

    # --- 企業基本情報 ---
    data.company.name = _safe_str(d.get("company_name"))
    data.company.representative = _safe_str(d.get("representative"))
    data.company.prefecture = _safe_str(d.get("prefecture"))
    data.company.address = _safe_str(d.get("address"))
    data.company.postal_code = _safe_str(d.get("postal_code"))
    data.company.phone = _safe_str(d.get("phone"))
    data.company.established_date = _safe_str(d.get("established_date"))
    data.company.capital = _safe_int(d.get("capital"))
    data.company.industry = _safe_str(d.get("industry"))
    data.company.business_description = _safe_str(d.get("business_description"))
    data.company.employee_count = _safe_int(d.get("employee_count"))
    data.company.officer_count = max(_safe_int(d.get("officer_count"), 1), 1)
    data.company.url = _safe_str(d.get("url"))

    # --- 人手不足情報 ---
    data.labor_shortage.shortage_tasks = _safe_str(d.get("shortage_tasks"))
    data.labor_shortage.recruitment_period = _safe_str(d.get("recruitment_period"))
    data.labor_shortage.applications = _safe_int(d.get("applications"))
    data.labor_shortage.hired = _safe_int(d.get("hired"))
    data.labor_shortage.overtime_hours = _safe_float(d.get("overtime_hours"))
    data.labor_shortage.current_workers = _safe_int(d.get("current_workers"))
    data.labor_shortage.desired_workers = _safe_int(d.get("desired_workers"))
    data.labor_shortage.job_openings_ratio = _safe_float(d.get("job_openings_ratio"))

    # --- 省力化効果 ---
    data.labor_saving.target_tasks = _safe_str(d.get("target_tasks"))
    data.labor_saving.current_hours = _safe_float(d.get("current_hours"))
    data.labor_saving.target_hours = _safe_float(d.get("target_hours"))
    if data.labor_saving.current_hours > 0:
        data.labor_saving.reduction_hours = max(
            0, data.labor_saving.current_hours - data.labor_saving.target_hours
        )
        data.labor_saving.reduction_rate = max(
            0,
            (data.labor_saving.reduction_hours / data.labor_saving.current_hours) * 100,
        )

    # --- 導入設備 ---
    data.equipment.name = _safe_str(d.get("equipment_name"))
    data.equipment.category = _safe_str(d.get("equipment_category"))
    data.equipment.manufacturer = _safe_str(d.get("manufacturer"))
    data.equipment.model = _safe_str(d.get("model"))
    data.equipment.quantity = max(_safe_int(d.get("quantity"), 1), 1)
    data.equipment.total_price = _safe_int(d.get("total_price"))
    data.equipment.vendor = _safe_str(d.get("vendor"))
    data.equipment.features = _safe_str(d.get("features"))
    data.equipment.catalog_number = _safe_str(d.get("catalog_number"))

    # --- 資金調達 ---
    data.funding.total_investment = _safe_int(d.get("total_investment"))
    data.funding.subsidy_amount = _safe_int(d.get("subsidy_amount"))
    data.funding.self_funding = _safe_int(d.get("self_funding"))
    data.funding.implementation_manager = _safe_str(d.get("implementation_manager"))
    data.funding.implementation_period = _safe_str(d.get("implementation_period"))
    data.funding.bank_name = _safe_str(d.get("bank_name"))

    # --- 財務情報 ---
    revenue = _safe_int(d.get("revenue"))
    gross_profit = _safe_int(d.get("gross_profit"))
    operating_profit = _safe_int(d.get("operating_profit"))

    if revenue > 0:
        data.company.revenue_2024 = revenue
        data.company.revenue_2023 = int(revenue / Config.GROWTH_RATE)
        data.company.revenue_2022 = int(revenue / Config.GROWTH_RATE / Config.GROWTH_RATE)
    if gross_profit > 0:
        data.company.gross_profit_2024 = gross_profit
        data.company.gross_profit_2023 = int(gross_profit / Config.GROWTH_RATE)
        data.company.gross_profit_2022 = int(gross_profit / Config.GROWTH_RATE / Config.GROWTH_RATE)
    if operating_profit > 0:
        data.company.operating_profit_2024 = operating_profit
        data.company.operating_profit_2023 = int(operating_profit / Config.PROFIT_GROWTH_RATE)
        data.company.operating_profit_2022 = int(operating_profit / Config.PROFIT_GROWTH_RATE / Config.PROFIT_GROWTH_RATE)

    data.company.labor_cost = _safe_int(d.get("labor_cost"))
    data.company.depreciation = _safe_int(d.get("depreciation"))
    data.company.total_salary = _safe_int(d.get("total_salary"))

    # --- 役員 ---
    for o in d.get("officers", []):
        if isinstance(o, dict) and o.get("name"):
            data.officers.append(OfficerInfo(
                name=_safe_str(o.get("name")),
                position=_safe_str(o.get("position", "役員")),
                birth_date=_safe_str(o.get("birth_date")),
            ))
    if data.officers:
        data.company.officer_count = len(data.officers)

    # --- 従業員 ---
    for e in d.get("employees", []):
        if isinstance(e, dict) and e.get("name"):
            data.employees.append(EmployeeInfo(
                name=_safe_str(e.get("name")),
                birth_date=_safe_str(e.get("birth_date")),
                hire_date=_safe_str(e.get("hire_date")),
            ))

    # --- 株主 ---
    for s in d.get("shareholders", []):
        if isinstance(s, dict) and s.get("name"):
            data.shareholders.append(ShareholderInfo(
                name=_safe_str(s.get("name")),
                shares=_safe_int(s.get("shares")),
            ))

    # --- 賃上げ・活用計画 ---
    data.wage_increase_rate = _safe_float(d.get("wage_increase_rate"))
    data.wage_increase_target = _safe_str(d.get("wage_increase_target"))
    data.wage_increase_timing = _safe_str(d.get("wage_increase_timing"))
    data.motivation_background = _safe_str(d.get("motivation_background"))
    data.time_utilization_plan = _safe_str(d.get("time_utilization_plan"))

    # --- フォールバック計算（read_hearing_sheet と同じロジック） ---
    if data.funding.total_investment == 0 and data.equipment.total_price > 0:
        data.funding.total_investment = data.equipment.total_price
    if data.funding.subsidy_amount == 0 and data.funding.total_investment > 0:
        data.funding.subsidy_amount = int(data.funding.total_investment * 0.5)
        print(f"  ⚠️ subsidy_amount=0 → total_investment×0.5で推計: {data.funding.subsidy_amount:,}円")
    if data.funding.self_funding == 0 and data.funding.total_investment > 0:
        data.funding.self_funding = data.funding.total_investment - data.funding.subsidy_amount
        print(f"  ⚠️ self_funding=0 → 差額で推計: {data.funding.self_funding:,}円")

    return data


# =============================================================================
# Excel書き出し（11シート）
# =============================================================================

def write_hearing_excel(data: HearingData, output_path: str) -> str:
    """HearingDataから11シートのExcelを生成する。

    シート名・ラベル文字列は read_hearing_sheet() の find_value() が
    検索するラベルと完全一致させ、ラウンドトリップを保証する。
    """
    wb = openpyxl.Workbook()

    # デフォルトシートを削除
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    # ----- 1. 企業基本情報 -----
    ws = wb.create_sheet("1_企業基本情報")
    labels_1 = [
        ("会社名", data.company.name),
        ("代表者名", data.company.representative),
        ("都道府県", data.company.prefecture),
        ("市区町村", data.company.address),
        ("郵便番号", data.company.postal_code),
        ("電話番号", data.company.phone),
        ("設立", data.company.established_date),
        ("資本金", data.company.capital),
        ("業種", data.company.industry),
        ("事業内容", data.company.business_description),
        ("従業員数", data.company.employee_count),
        ("URL", data.company.url),
    ]
    for i, (label, val) in enumerate(labels_1, start=1):
        ws.cell(row=i, column=1, value=label)
        ws.cell(row=i, column=2, value=val)

    # ----- 2. 人手不足の実態 -----
    ws = wb.create_sheet("2_人手不足")
    labels_2 = [
        ("困っている作業", data.labor_shortage.shortage_tasks),
        ("求人を出している期間", data.labor_shortage.recruitment_period),
        ("応募", data.labor_shortage.applications),
        ("採用できた", data.labor_shortage.hired),
        ("残業", data.labor_shortage.overtime_hours),
        ("何人でやっています", data.labor_shortage.current_workers),
        ("何人いれば", data.labor_shortage.desired_workers),
        ("有効求人倍率", data.labor_shortage.job_openings_ratio),
    ]
    for i, (label, val) in enumerate(labels_2, start=1):
        ws.cell(row=i, column=1, value=label)
        ws.cell(row=i, column=2, value=val)

    # ----- 3. なぜ今必要か -----
    ws = wb.create_sheet("3_なぜ今")
    ws.cell(row=1, column=1, value="背景")
    ws.cell(row=1, column=2, value=data.motivation_background)

    # ----- 4. 省力化効果 -----
    ws = wb.create_sheet("4_省力化効果")
    labels_4 = [
        ("対象となる作業", data.labor_saving.target_tasks),
        ("導入前", data.labor_saving.current_hours),
        ("導入後", data.labor_saving.target_hours),
    ]
    for i, (label, val) in enumerate(labels_4, start=1):
        ws.cell(row=i, column=1, value=label)
        ws.cell(row=i, column=2, value=val)

    # ----- 5. 導入設備 -----
    ws = wb.create_sheet("5_導入設備")
    labels_5 = [
        ("設備の名前", data.equipment.name),
        ("設備カテゴリ", data.equipment.category),
        ("メーカー", data.equipment.manufacturer),
        ("型番", data.equipment.model),
        ("数量", data.equipment.quantity),
        ("購入先", data.equipment.vendor),
        ("金額", data.equipment.total_price),
        ("カスタマイズ", data.equipment.features),
        ("カタログ", data.equipment.catalog_number),
    ]
    for i, (label, val) in enumerate(labels_5, start=1):
        ws.cell(row=i, column=1, value=label)
        ws.cell(row=i, column=2, value=val)

    # ----- 6. 効果の活用計画 -----
    ws = wb.create_sheet("6_効果の活用")
    ws.cell(row=1, column=1, value="活用")
    ws.cell(row=1, column=2, value=data.time_utilization_plan)

    # ----- 7. 賃上げ計画 -----
    ws = wb.create_sheet("7_賃上げ")
    labels_7 = [
        ("賃上げ率", data.wage_increase_rate),
        ("対象者", data.wage_increase_target),
        ("実施時期", data.wage_increase_timing),
    ]
    for i, (label, val) in enumerate(labels_7, start=1):
        ws.cell(row=i, column=1, value=label)
        ws.cell(row=i, column=2, value=val)

    # ----- 8. 資金調達・体制 -----
    ws = wb.create_sheet("8_資金調達")
    labels_8 = [
        ("投資総額", data.funding.total_investment),
        ("補助金申請額", data.funding.subsidy_amount),
        ("自己資金", data.funding.self_funding),
        ("責任者", data.funding.implementation_manager),
        ("実施期間", data.funding.implementation_period),
        ("銀行", data.funding.bank_name),
    ]
    for i, (label, val) in enumerate(labels_8, start=1):
        ws.cell(row=i, column=1, value=label)
        ws.cell(row=i, column=2, value=val)

    # ----- 9. 従業員情報 -----
    ws = wb.create_sheet("9_従業員情報")
    ws.cell(row=1, column=1, value="No.")
    ws.cell(row=1, column=2, value="氏名")
    ws.cell(row=1, column=3, value="生年月日")
    ws.cell(row=1, column=4, value="入社日")
    for i, emp in enumerate(data.employees, start=2):
        ws.cell(row=i, column=1, value=i - 1)
        ws.cell(row=i, column=2, value=emp.name)
        ws.cell(row=i, column=3, value=emp.birth_date)
        ws.cell(row=i, column=4, value=emp.hire_date)

    # ----- 10. 役員・株主情報 -----
    ws = wb.create_sheet("10_役員_株主")
    # 役員セクション
    ws.cell(row=1, column=1, value="役員情報")
    ws.cell(row=2, column=1, value="No.")
    ws.cell(row=2, column=2, value="氏名")
    ws.cell(row=2, column=3, value="役職")
    ws.cell(row=2, column=4, value="生年月日")
    row = 3
    for i, off in enumerate(data.officers, start=1):
        ws.cell(row=row, column=1, value=i)
        ws.cell(row=row, column=2, value=off.name)
        ws.cell(row=row, column=3, value=off.position)
        ws.cell(row=row, column=4, value=off.birth_date)
        row += 1

    # 株主セクション
    row += 1
    ws.cell(row=row, column=1, value="株主情報")
    row += 1
    ws.cell(row=row, column=1, value="No.")
    ws.cell(row=row, column=2, value="株主名")
    ws.cell(row=row, column=3, value="株数")
    row += 1
    for i, sh in enumerate(data.shareholders, start=1):
        ws.cell(row=row, column=1, value=i)
        ws.cell(row=row, column=2, value=sh.name)
        ws.cell(row=row, column=3, value=sh.shares)
        row += 1

    # ----- 11. 財務情報 -----
    ws = wb.create_sheet("11_財務情報")
    labels_11 = [
        ("売上高", data.company.revenue_2024),
        ("売上総利益", data.company.gross_profit_2024),
        ("営業利益", data.company.operating_profit_2024),
        ("人件費", data.company.labor_cost),
        ("減価償却費", data.company.depreciation),
        ("給与支給総額", data.company.total_salary),
    ]
    for i, (label, val) in enumerate(labels_11, start=1):
        ws.cell(row=i, column=1, value=label)
        ws.cell(row=i, column=2, value=val)

    wb.save(output_path)
    wb.close()
    print(f"  ✅ ヒアリングシート生成: {output_path}")
    return output_path


# =============================================================================
# メインパイプライン
# =============================================================================

def transcription_to_hearing(
    input_path: str,
    output_path: str,
    api_key: str,
    model: str = "claude-sonnet-4-20250514",
) -> tuple:
    """議事録テキストからヒアリングシートExcelを生成する。

    Returns:
        (HearingData, ExtractionResult, output_path)
    """
    print("=" * 60)
    print("議事録 → ヒアリングシート自動生成")
    print("=" * 60)

    # 1. テキスト読み込み
    print(f"\n📄 議事録読み込み: {input_path}")
    with open(input_path, "r", encoding="utf-8") as f:
        text = f.read()
    print(f"  文字数: {len(text):,}")

    # 2. Claude APIで抽出
    print(f"\n🤖 Claude APIによるデータ抽出 (model: {model})")
    raw = extract_from_transcription(text, api_key, model)

    # 3. バリデーション
    print("\n🔍 バリデーション")
    extraction_result = validate_extracted_data(raw)
    for w in extraction_result.warnings:
        print(f"  ⚠️ {w}")

    # 4. HearingDataに変換
    print("\n🔄 HearingData変換")
    hearing_data = build_hearing_data(extraction_result)

    # 5. Excel書き出し
    print(f"\n📝 Excel書き出し: {output_path}")
    write_hearing_excel(hearing_data, output_path)

    print("\n" + "=" * 60)
    print(f"  企業名: {hearing_data.company.name}")
    print(f"  業種: {hearing_data.company.industry}")
    print(f"  設備: {hearing_data.equipment.name}")
    print(f"  投資額: {hearing_data.equipment.total_price:,}円")
    if extraction_result.missing_critical:
        print(f"  ⚠️ 要確認: {', '.join(extraction_result.missing_critical)}")
    print("=" * 60)

    return hearing_data, extraction_result, output_path


# =============================================================================
# CLI
# =============================================================================

def main():
    parser = argparse.ArgumentParser(
        description="議事録テキスト → ヒアリングシート自動生成"
    )
    parser.add_argument("--input", "-i", required=True, help="議事録テキストファイルのパス")
    parser.add_argument("--output", "-o", required=True, help="出力するExcelファイルのパス")
    parser.add_argument(
        "--generate", "-g", action="store_true",
        help="ヒアリングシート生成後、書類生成まで一気通貫で実行",
    )
    parser.add_argument("--model", default="claude-sonnet-4-20250514", help="Claude APIモデル")
    parser.add_argument("--template-dir", "-t", default="./templates", help="テンプレートディレクトリ（--generate時）")
    parser.add_argument("--output-dir", default="./output", help="書類出力ディレクトリ（--generate時）")
    parser.add_argument("--no-diagrams", action="store_true", help="図解生成をスキップ（--generate時）")
    args = parser.parse_args()

    # APIキー取得
    api_key = os.environ.get("ANTHROPIC_API_KEY", "")
    if not api_key:
        print("❌ ANTHROPIC_API_KEY 環境変数を設定してください。")
        sys.exit(1)

    # パイプライン実行
    hearing_data, extraction_result, excel_path = transcription_to_hearing(
        input_path=args.input,
        output_path=args.output,
        api_key=api_key,
        model=args.model,
    )

    # --generate: 書類生成まで続行
    if args.generate:
        print("\n🚀 書類生成パイプラインに接続...")
        from main import (
            generate_diagrams,
            generate_with_auto_fix,
        )

        template_dir = Path(args.template_dir)
        output_dir = Path(args.output_dir)
        output_dir.mkdir(exist_ok=True, parents=True)

        # read_hearing_sheet でラウンドトリップ確認
        hearing_data = read_hearing_sheet(excel_path)

        diagrams = {}
        if not args.no_diagrams:
            diagrams = generate_diagrams(hearing_data, str(output_dir))

        result = generate_with_auto_fix(
            data=hearing_data,
            output_dir=str(output_dir),
            template_dir=template_dir,
            diagrams=diagrams,
            target_score=85,
            max_iterations=5,
            skip_diagrams=args.no_diagrams,
        )
        print(f"\n✅ 書類生成完了 — スコア: {result['score']}/100")
        print(f"📁 出力先: {output_dir}")


if __name__ == "__main__":
    main()
