#!/usr/bin/env python3
"""事業計画書Part3 Excel生成"""

import os
import shutil
from pathlib import Path

import openpyxl

from models import HearingData
from config import Config
from hearing_reader import _find_sheet_in_workbook


def generate_business_plan_3(data: HearingData, output_dir: str, template_path: Path):
    """
    事業計画書その3を生成（openpyxl方式）
    """
    print("\n📊 事業計画書（その3）を生成中...")

    c, e, f = data.company, data.equipment, data.funding

    # 出力先にコピー
    output_path = Path(output_dir) / "事業計画書_その3_完成版.xlsx"
    shutil.copy(template_path, output_path)
    os.chmod(output_path, 0o644)

    try:
        wb = openpyxl.load_workbook(output_path)

        # Phase 6: _find_sheet_in_workbook でシート検索
        # === 別紙1（省力化計算シート）===
        ws1 = _find_sheet_in_workbook(wb, ["別紙1", "省力化"])
        if ws1:
            print(f"    📋 別紙1: {ws1.title}")

            # 導入前工程（C11〜C16, D11〜D16）
            for i, p in enumerate(data.before_processes[:6]):
                ws1[f'C{11+i}'] = p.name
                ws1[f'D{11+i}'] = p.time_minutes

            # 導入後工程（I11〜I16, J11〜J16）
            for i, p in enumerate(data.after_processes[:6]):
                ws1[f'I{11+i}'] = p.name
                ws1[f'J{11+i}'] = p.time_minutes

            print("      ✅ 工程データ入力完了")

        # === 別紙3（投資回収期間計算シート）===
        ws3 = _find_sheet_in_workbook(wb, ["別紙3", "投資回収"])
        if ws3:
            print(f"    📋 別紙3: {ws3.title}")

            # Phase 2: Config参照
            ws3['C6'] = f.total_investment  # 投資総額
            ws3['H6'] = Config.WORKING_DAYS_PER_YEAR  # 年間稼働日数
            ws3['J6'] = Config.HOURLY_WAGE  # 人件費単価
            ws3['L6'] = int(e.total_price / Config.DEPRECIATION_YEARS)  # 減価償却費

            print("      ✅ 投資回収データ入力完了")

        # === 参考書式（事業計画目標値算出シート）===
        ws_ref = _find_sheet_in_workbook(wb, ["参考書式", "目標値"])
        if ws_ref:
            print(f"    📋 参考書式: {ws_ref.title}")

            # --- ラベル行を動的検索 ---
            def find_row_by_label(ws, keywords, search_cols=('A', 'B', 'C', 'D'), max_row=60):
                """シート内でキーワードに一致するラベル行を見つける"""
                for row in range(1, max_row + 1):
                    for col_letter in search_cols:
                        val = ws[f'{col_letter}{row}'].value
                        if val:
                            for kw in keywords:
                                if kw in str(val):
                                    return row
                return None

            # 行番号を検索（フォールバック付き）
            row_revenue = find_row_by_label(ws_ref, ["売上高"]) or 26
            row_operating_profit = find_row_by_label(ws_ref, ["営業利益"])
            row_labor_cost = find_row_by_label(ws_ref, ["人件費"])
            row_depreciation = find_row_by_label(ws_ref, ["減価償却費"])
            row_added_value = find_row_by_label(ws_ref, ["付加価値額"])
            row_officers = find_row_by_label(ws_ref, ["役員数"]) or 37
            row_employees = find_row_by_label(ws_ref, ["従業員数"]) or 38
            row_salary_total = find_row_by_label(ws_ref, ["給与支給総額"]) or 44
            row_salary_employees = find_row_by_label(ws_ref, ["給与対象"]) or 45

            base_revenue = c.revenue_2024
            base_op_profit = c.operating_profit_2024

            # 人件費: 決算書PDFから取得 or 売上高×労働分配率で推計
            base_labor_cost = c.labor_cost if c.labor_cost > 0 else int(base_revenue * Config.LABOR_COST_RATIO)
            # 減価償却費: 決算書PDFから取得 or 設備投資額÷耐用年数で推計
            base_depreciation = c.depreciation if c.depreciation > 0 else int(e.total_price / Config.DEPRECIATION_YEARS)
            # 給与支給総額: 決算書PDFから取得 or 売上高×給与比率で推計
            base_salary = c.total_salary if c.total_salary > 0 else int(base_revenue * Config.SALARY_RATIO)
            # 付加価値額 = 営業利益 + 人件費 + 減価償却費
            base_added_value = base_op_profit + base_labor_cost + base_depreciation

            # E列=基準, G〜K列=1〜5年目
            cols = ['E', 'G', 'H', 'I', 'J', 'K']

            for i, col in enumerate(cols):
                growth = Config.GROWTH_RATE ** i
                salary_growth = Config.SALARY_GROWTH_RATE ** i

                # 売上高
                ws_ref[f'{col}{row_revenue}'] = int(base_revenue * growth)
                # 営業利益
                if row_operating_profit:
                    ws_ref[f'{col}{row_operating_profit}'] = int(base_op_profit * growth)
                # 人件費
                if row_labor_cost:
                    ws_ref[f'{col}{row_labor_cost}'] = int(base_labor_cost * salary_growth)
                # 減価償却費
                if row_depreciation:
                    ws_ref[f'{col}{row_depreciation}'] = int(base_depreciation)
                # 付加価値額
                if row_added_value:
                    av_op = int(base_op_profit * growth)
                    av_lc = int(base_labor_cost * salary_growth)
                    av_dep = int(base_depreciation)
                    ws_ref[f'{col}{row_added_value}'] = av_op + av_lc + av_dep
                # 役員数
                ws_ref[f'{col}{row_officers}'] = c.officer_count
                # 従業員数
                ws_ref[f'{col}{row_employees}'] = c.employee_count
                # 給与支給総額（年率2.5%成長）
                ws_ref[f'{col}{row_salary_total}'] = int(base_salary * salary_growth)
                # 給与対象従業員数
                ws_ref[f'{col}{row_salary_employees}'] = c.employee_count

            # 成長率の確認ログ
            year5_added_value = int(base_op_profit * Config.GROWTH_RATE**5) + int(base_labor_cost * Config.SALARY_GROWTH_RATE**5) + int(base_depreciation)
            if base_added_value > 0:
                av_annual_growth = ((year5_added_value / base_added_value) ** (1/5) - 1) * 100
                print(f"      📊 付加価値額: 基準{base_added_value:,}円 → 5年目{year5_added_value:,}円（年率{av_annual_growth:.1f}%）")
            year5_salary = int(base_salary * Config.SALARY_GROWTH_RATE**5)
            if base_salary > 0:
                sal_annual_growth = ((year5_salary / base_salary) ** (1/5) - 1) * 100
                print(f"      📊 給与支給総額: 基準{base_salary:,}円 → 5年目{year5_salary:,}円（年率{sal_annual_growth:.1f}%）")

            print("      ✅ 目標値データ入力完了")

        # 保存
        wb.save(output_path)
        wb.close()

        print(f"  ✅ 保存完了: {output_path.name}")
        print("    ⚠️ 注意: ファイルが開けない場合は手動でコピーが必要です")

    except Exception as ex:
        print(f"    ⚠️ openpyxlエラー: {ex}")
        print("    📝 2ファイル方式にフォールバック...")
        # フォールバック: 2ファイル方式
        from openpyxl import Workbook
        data_file = Path(output_dir) / "事業計画書_その3_入力データ.xlsx"
        wb_new = Workbook()
        ws1 = wb_new.active
        ws1.title = "別紙1_工程データ"
        for i, p in enumerate(data.before_processes):
            ws1.cell(row=2+i, column=1, value=p.name)
            ws1.cell(row=2+i, column=2, value=p.time_minutes)
        for i, p in enumerate(data.after_processes):
            ws1.cell(row=2+i, column=4, value=p.name)
            ws1.cell(row=2+i, column=5, value=p.time_minutes)
        wb_new.save(data_file)
        print(f"    ✅ 入力データ: {data_file.name}")
