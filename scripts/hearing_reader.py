#!/usr/bin/env python3
"""ヒアリングシート読み込み"""

from typing import List, Tuple

import openpyxl

from models import (
    HearingData, CompanyInfo, LaborShortageInfo, LaborSavingInfo,
    EquipmentInfo, FundingInfo, OfficerInfo, EmployeeInfo,
    ShareholderInfo, WorkProcess,
)
from config import Config
from process_templates import generate_processes


def _split_name(name: str) -> Tuple[str, str]:
    """姓名を分割する（全角スペース対応）（Phase 6）"""
    name = str(name).strip()
    # 全角スペース → 半角スペースに正規化してから分割
    normalized = name.replace('\u3000', ' ')
    parts = normalized.split()
    if len(parts) >= 2:
        return parts[0], ' '.join(parts[1:])
    return name, ''


def _find_sheet_in_workbook(wb, patterns: list):
    """ファジーマッチングでワークブック内のシートを検索する（Phase 6）"""
    for name in wb.sheetnames:
        for p in patterns:
            if p in name:
                return wb[name]
    return None


def read_hearing_sheet(file_path: str) -> HearingData:
    """ヒアリングシートから全データを読み込む"""
    print(f"📖 ヒアリングシート読み込み中: {file_path}")

    # Phase 1: リソースリーク防止 — try-except-finally で wb.close() を保証
    wb = None
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
    except FileNotFoundError:
        print(f"  ❌ ファイルが見つかりません: {file_path}")
        raise
    except Exception as ex:
        print(f"  ❌ ファイルオープンエラー: {ex}")
        raise

    data = HearingData()

    try:
        # ヘルパー関数
        def find_value(ws, labels, offset=1):
            """ラベルに対応する値を検索"""
            if isinstance(labels, str):
                labels = [labels]
            for row in range(1, 50):
                for col in range(1, 10):
                    val = ws.cell(row=row, column=col).value
                    if val:
                        for label in labels:
                            if label in str(val):
                                result = ws.cell(row=row, column=col + offset).value
                                # Phase 1: None ガード
                                return result if result is not None else ""
            return ""

        def find_int(ws, labels, offset=1, default=0):
            """整数値を検索"""
            val = find_value(ws, labels, offset)
            if not val:
                return default
            try:
                return int(float(str(val).replace(",", "").replace("円", "").replace("人", "").replace("名", "")))
            except (ValueError, TypeError):
                return default

        def find_float(ws, labels, offset=1, default=0.0):
            """浮動小数点値を検索"""
            val = find_value(ws, labels, offset)
            if not val:
                return default
            try:
                return float(str(val).replace(",", "").replace("%", "").replace("時間", ""))
            except (ValueError, TypeError):
                return default

        def find_sheet(patterns):
            """パターンに一致するシートを検索（Phase 6: ファジーマッチング関数を使用）"""
            return _find_sheet_in_workbook(wb, patterns)

        # ----- 1. 企業基本情報 -----
        ws = find_sheet(["企業基本情報", "1_"])
        if ws:
            data.company.name = str(find_value(ws, ["会社名", "法人名"]))
            data.company.representative = str(find_value(ws, ["代表者名", "代表取締役"]))
            data.company.prefecture = str(find_value(ws, ["都道府県"]))
            data.company.address = str(find_value(ws, ["市区町村"]))
            data.company.phone = str(find_value(ws, ["電話番号"]))
            data.company.industry = str(find_value(ws, ["業種"]))
            data.company.business_description = str(find_value(ws, ["事業内容", "主な事業内容"]))
            data.company.employee_count = find_int(ws, ["従業員数", "常勤"])
            data.company.established_date = str(find_value(ws, ["設立"]))
            data.company.capital = find_int(ws, ["資本金"])
            data.company.url = str(find_value(ws, ["URL", "ホームページ"]))

        # ----- 2. 人手不足の実態 -----
        ws = find_sheet(["人手不足", "2_"])
        if ws:
            data.labor_shortage.shortage_tasks = str(find_value(ws, ["困っている作業", "人が足りなくて"]))
            data.labor_shortage.recruitment_period = str(find_value(ws, ["求人を出している期間"]))
            data.labor_shortage.applications = find_int(ws, ["応募"])
            data.labor_shortage.hired = find_int(ws, ["採用できた"])
            data.labor_shortage.overtime_hours = find_float(ws, ["残業"])
            data.labor_shortage.current_workers = find_int(ws, ["何人でやっています"])
            data.labor_shortage.desired_workers = find_int(ws, ["何人いれば"])
            data.labor_shortage.job_openings_ratio = find_float(ws, ["有効求人倍率"])

        # ----- 3. 省力化効果 -----
        ws = find_sheet(["省力化効果", "4_"])
        if ws:
            data.labor_saving.target_tasks = str(find_value(ws, ["対象となる作業"]))
            data.labor_saving.current_hours = find_float(ws, ["導入前", "1日に何時間", "何時間その作業", "現在の作業時間", "今かかっている"])
            data.labor_saving.target_hours = find_float(ws, ["導入後", "導入したら何時間", "目標時間", "何時間になりそう", "短縮後"])
            if data.labor_saving.current_hours > 0:
                # Phase 1: マイナス削減率防止
                data.labor_saving.reduction_hours = max(0, data.labor_saving.current_hours - data.labor_saving.target_hours)
                data.labor_saving.reduction_rate = max(0, (data.labor_saving.reduction_hours / data.labor_saving.current_hours) * 100)
                if data.labor_saving.target_hours > data.labor_saving.current_hours:
                    print(f"  ⚠️ 警告: 導入後の時間({data.labor_saving.target_hours}h)が導入前({data.labor_saving.current_hours}h)より大きいです。削減率を0%にしました。")

        # ----- 4. 導入設備 -----
        ws = find_sheet(["5_導入", "導入設備"])
        if ws:
            data.equipment.name = str(find_value(ws, ["設備の名前", "設備名"]))
            data.equipment.category = str(find_value(ws, ["設備カテゴリ", "何をするもの"]))
            data.equipment.manufacturer = str(find_value(ws, ["メーカー"]))
            data.equipment.model = str(find_value(ws, ["型番"]))
            data.equipment.quantity = find_int(ws, ["数量"], default=1)
            data.equipment.vendor = str(find_value(ws, ["購入先", "ベンダー", "どこから買"]))
            data.equipment.total_price = find_int(ws, ["金額", "税抜", "いくら"])
            data.equipment.features = str(find_value(ws, ["カスタマイズ", "特徴"]))
            data.equipment.catalog_number = str(find_value(ws, ["カタログ", "登録番号"]))

        # ----- 5. 資金調達・体制 -----
        ws = find_sheet(["資金調達", "体制", "8_"])
        if ws:
            data.funding.total_investment = find_int(ws, ["投資総額", "設備の金額"])
            data.funding.subsidy_amount = find_int(ws, ["補助金申請額", "補助金"])
            data.funding.self_funding = find_int(ws, ["自己資金"])
            data.funding.implementation_manager = str(find_value(ws, ["責任者"]))
            data.funding.implementation_period = str(find_value(ws, ["実施期間", "導入時期"]))
            data.funding.bank_name = str(find_value(ws, ["銀行", "借入先", "取引銀行"]))

        if data.funding.total_investment == 0:
            data.funding.total_investment = data.equipment.total_price

        # ----- 6. 財務情報 -----
        ws = find_sheet(["11_財務", "財務情報"])
        if ws:
            base_revenue = find_int(ws, ["売上高"], default=Config.DEFAULT_REVENUE)
            base_gross_profit = find_int(ws, ["売上総利益", "粗利"], default=int(base_revenue * 0.7))
            base_profit = find_int(ws, ["営業利益"], default=Config.DEFAULT_PROFIT)

            # 基準年度から3年分を推計（Phase 2: Config参照）
            data.company.revenue_2024 = base_revenue
            data.company.revenue_2023 = int(base_revenue / Config.GROWTH_RATE)
            data.company.revenue_2022 = int(base_revenue / Config.GROWTH_RATE / Config.GROWTH_RATE)

            data.company.gross_profit_2024 = base_gross_profit
            data.company.gross_profit_2023 = int(base_gross_profit / Config.GROWTH_RATE)
            data.company.gross_profit_2022 = int(base_gross_profit / Config.GROWTH_RATE / Config.GROWTH_RATE)

            data.company.operating_profit_2024 = base_profit
            data.company.operating_profit_2023 = int(base_profit / Config.PROFIT_GROWTH_RATE)
            data.company.operating_profit_2022 = int(base_profit / Config.PROFIT_GROWTH_RATE / Config.PROFIT_GROWTH_RATE)
        else:
            # デフォルト値
            data.company.revenue_2022, data.company.revenue_2023, data.company.revenue_2024 = 47000000, 49000000, 50000000
            data.company.gross_profit_2022, data.company.gross_profit_2023, data.company.gross_profit_2024 = 33000000, 34000000, 35000000
            data.company.operating_profit_2022, data.company.operating_profit_2023, data.company.operating_profit_2024 = 6000000, 6500000, 7000000

        # ----- 7. 役員情報（株主セクションの手前まで）-----
        ws = find_sheet(["10_役員", "役員_株主"])
        if ws:
            in_officer_section = False
            for row in range(1, ws.max_row + 1):
                col1 = ws.cell(row=row, column=1).value
                name = ws.cell(row=row, column=2).value

                if col1 and "役員情報" in str(col1):
                    in_officer_section = True
                    continue
                if col1 and "株主情報" in str(col1):
                    break

                if in_officer_section and name and str(name).strip() and "氏名" not in str(name):
                    data.officers.append(OfficerInfo(
                        name=str(name),
                        position=str(ws.cell(row=row, column=3).value or "役員"),
                        birth_date=str(ws.cell(row=row, column=4).value or "")
                    ))
            data.company.officer_count = max(len(data.officers), 1)

        # ----- 8. 従業員情報 -----
        ws = find_sheet(["9_従業員", "従業員情報"])
        if ws:
            for row in range(2, ws.max_row + 1):
                name = ws.cell(row=row, column=2).value
                if name and str(name).strip() and "氏名" not in str(name):
                    data.employees.append(EmployeeInfo(
                        name=str(name),
                        birth_date=str(ws.cell(row=row, column=3).value or ""),
                        hire_date=str(ws.cell(row=row, column=4).value or "")
                    ))

        # ----- 9. 株主情報 -----
        ws = find_sheet(["10_役員", "役員_株主"])
        if ws:
            in_shareholder_section = False
            for row in range(1, ws.max_row + 1):
                col1 = ws.cell(row=row, column=1).value
                name = ws.cell(row=row, column=2).value

                if col1 and "株主情報" in str(col1):
                    in_shareholder_section = True
                    continue

                if in_shareholder_section and name and str(name).strip() and "株主名" not in str(name):
                    try:
                        shares = int(ws.cell(row=row, column=3).value or 0)
                    except (ValueError, TypeError):
                        shares = 0
                    data.shareholders.append(ShareholderInfo(name=str(name), shares=shares))

        # ----- Phase 4: シート3（なぜ今必要か）-----
        ws = find_sheet(["なぜ今", "3_"])
        if ws:
            data.motivation_background = str(find_value(ws, ["背景", "なぜ今", "理由", "きっかけ"]))

        # ----- Phase 4: シート6（効果の活用計画）-----
        ws = find_sheet(["効果の活用", "6_"])
        if ws:
            data.time_utilization_plan = str(find_value(ws, ["活用", "浮いた時間", "時間の使い方"]))

        # ----- Phase 4: シート7（賃上げ計画）-----
        ws = find_sheet(["賃上げ", "7_"])
        if ws:
            data.wage_increase_rate = find_float(ws, ["賃上げ率", "引上げ率"])
            data.wage_increase_target = str(find_value(ws, ["対象者", "対象"]))
            data.wage_increase_timing = str(find_value(ws, ["実施時期", "いつから"]))

        # ----- 10. 工程データ生成 -----
        data.before_processes, data.after_processes = generate_processes(data)

    finally:
        # Phase 1: リソースリーク防止 — 必ず wb.close() を実行
        if wb is not None:
            wb.close()

    # フォールバック計算（0値の補完）
    if data.labor_saving.current_hours == 0 and data.before_processes:
        data.labor_saving.current_hours = sum(p.time_minutes for p in data.before_processes) / 60
        print(f"  ⚠️ current_hours=0 → before_processesから推計: {data.labor_saving.current_hours:.1f}h")
    if data.labor_saving.target_hours == 0 and data.after_processes:
        data.labor_saving.target_hours = sum(p.time_minutes for p in data.after_processes) / 60
        print(f"  ⚠️ target_hours=0 → after_processesから推計: {data.labor_saving.target_hours:.1f}h")
    if data.labor_saving.current_hours > 0 and data.labor_saving.target_hours > 0:
        data.labor_saving.reduction_hours = max(0, data.labor_saving.current_hours - data.labor_saving.target_hours)
        data.labor_saving.reduction_rate = max(0, (data.labor_saving.reduction_hours / data.labor_saving.current_hours) * 100)
    if data.funding.subsidy_amount == 0 and data.funding.total_investment > 0:
        data.funding.subsidy_amount = int(data.funding.total_investment * 0.5)
        print(f"  ⚠️ subsidy_amount=0 → total_investment×0.5で推計: {data.funding.subsidy_amount:,}円")
    if data.funding.self_funding == 0 and data.funding.total_investment > 0:
        data.funding.self_funding = data.funding.total_investment - data.funding.subsidy_amount
        print(f"  ⚠️ self_funding=0 → 差額で推計: {data.funding.self_funding:,}円")

    # 読み込み結果表示
    print(f"  ✅ 企業名: {data.company.name}")
    print(f"  ✅ 業種: {data.company.industry}")
    print(f"  ✅ 役員: {data.company.officer_count}名 / 従業員: {data.company.employee_count}名")
    print(f"  ✅ 設備: {data.equipment.name}")
    print(f"  ✅ 投資額: {data.equipment.total_price:,}円")
    print(f"  ✅ 補助金: {data.funding.subsidy_amount:,}円")
    print(f"  ✅ 削減率: {data.labor_saving.reduction_rate:.1f}%")
    print(f"  ✅ 売上高(2024): {data.company.revenue_2024:,}円")

    return data


def validate_hearing_data(data: HearingData) -> List[str]:
    """ヒアリングデータの必須フィールドを検証し、問題リストを返す"""
    issues = []
    if data.labor_saving.current_hours <= 0:
        issues.append("導入前の作業時間(current_hours)が0です")
    if data.labor_saving.target_hours <= 0:
        issues.append("導入後の作業時間(target_hours)が0です")
    if data.funding.subsidy_amount <= 0:
        issues.append("補助金申請額(subsidy_amount)が0です")
    if data.funding.total_investment <= 0:
        issues.append("投資総額(total_investment)が0です")
    if data.equipment.total_price <= 0:
        issues.append("設備価格(total_price)が0です")
    if not data.company.name or data.company.name.strip() == "":
        issues.append("企業名が空です")
    if data.company.employee_count <= 0:
        issues.append("従業員数(employee_count)が0です")
    return issues
